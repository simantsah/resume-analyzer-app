# Main Streamlit App
def main():
    st.set_page_config(page_title="Resume Analyzer", layout="wide", initial_sidebar_state="expanded")
    
    # Create tabs for different views
    tab1, tab2 = st.tabs(["Resume Analysis", "Candidates Dashboard"])
    
    with tab1:
        st.title("📝 Enhanced Resume Analyzer")
        st.markdown("Built with AI-powered skill matching and scoring")
        
        # Original app functionality
        main_analysis_tab()
    
    with tab2:
        st.title("📊 Candidates Dashboard")
        st.markdown("Overview of all analyzed candidates")
        
        dashboard_tab()

# Original analysis tab functionality
def main_analysis_tab():
    with st.sidebar:
        st.title("Scoring Algorithm")
        st.markdown("""
        ### Overall Score Formula
        - 40% × Relevancy Score (70% Strong + 30% Partial Matches)
        - 15% × Experience Score
        - 12% × Job Stability Score
        - 10% × College Rating
        - 10% × Leadership Score
        - 8% × International Experience
        - 5% × Competitor Experience
        
        ### Score Explanations
        - **Strong Matches**: Exact matching skills found in both JD and resume
        - **Partial Matches**: Related but different skills (e.g., PowerBI instead of Tableau)
        - **Relevancy Score**: Weighted combination of Strong (70%) and Partial (30%) matches
        - **Overall Weighted Score**: Combines relevancy with other factors using weights above
        
        ### Selection Categories
        - **Strong Fit (85-100) ✅**: Priority interview
        - **Good Fit (70-84) ✅**: Recommend interview
        - **Consider (55-69) 🤔**: Further screening needed
        - **Weak Fit (40-54) ⚠️**: Interview if candidate pool is limited
        - **Reject (0-39) ❌**: Does not meet minimum criteria
        """)
        
        # Add timer metrics display in sidebar
        with st.expander("⏱️ Performance Metrics", expanded=True):
            st.markdown("### Processing Times")
            st.markdown("Track how long each resume takes to process:")
            
            # Create placeholder metrics that will be updated during processing
            current_timer_container = st.empty()
            api_call_timer_container = st.empty()
            parsing_timer_container = st.empty()
            avg_timer_container = st.empty()
            total_timer_container = st.empty()
    
    try:
        load_dotenv()
        
        api_key = os.environ.get("GROQ_API_KEY")
        if not api_key:
            st.error("GROQ_API_KEY not found. Please set it in your environment or .env file.")
            st.info("You can get an API key from https://console.groq.com/")
            return
            
        client = initialize_groq_client()
        if not client:
            st.error("Failed to initialize Groq client. Please check your API key.")
            return
            
        uploaded_files = st.file_uploader("Upload resumes (PDF)", type=['pdf'], accept_multiple_files=True)
        job_description = st.text_area("Paste the job description here", height=200)
        
        # Create results_data in session state if it doesn't exist
        if 'results_data' not in st.session_state:
            st.session_state.results_data = []
        
        # Initialize timer metrics in session state
        if 'total_processing_time' not in st.session_state:
            st.session_state.total_processing_time = 0
        if 'processed_count' not in st.session_state:
            st.session_state.processed_count = 0
        if 'total_api_time' not in st.session_state:
            st.session_state.total_api_time = 0
        if 'total_parsing_time' not in st.session_state:
            st.session_state.total_parsing_time = 0
        if 'total_extraction_time' not in st.session_state:
            st.session_state.total_extraction_time = 0
        
        if uploaded_files and job_description:
            if st.button("Analyze All Resumes"):
                progress_bar = st.progress(0)
                total_files = len(uploaded_files)
                
                # Clear previous results when starting a new batch
                st.session_state.results_data = []
                
                # Start batch timing
                batch_start_time = time.time()
                
                for i, uploaded_file in enumerate(uploaded_files):
                    st.subheader(f"Resume: {uploaded_file.name}")
                    
                    # Start timer for individual resume
                    resume_start_time = time.time()
                    current_timer_container.metric("⏱️ Current Resume", "Processing...")
                    
                    with st.spinner(f"Analyzing {uploaded_file.name}..."):
                        # Time the PDF extraction
                        extraction_start = time.time()
                        resume_text = extract_text_from_pdf(uploaded_file)
                        extraction_time = time.time() - extraction_start
                        st.session_state.total_extraction_time += extraction_time
                        
                        if resume_text:
                            # Time the AI analysis (API call)
                            api_call_start = time.time()
                            analysis = analyze_resume(client, resume_text, job_description)
                            api_call_time = time.time() - api_call_start
                            
                            # Extract API call time from embedded data in analysis
                            api_time_match = re.search(r'API call time: (\d+\.\d+)', analysis) if analysis else None
                            if api_time_match:
                                api_call_time = float(api_time_match.group(1))
                            
                            st.session_state.total_api_time += api_call_time
                            api_call_timer_container.metric("⏱️ API Call", f"{api_call_time:.2f} seconds")
                            
                            if analysis:
                                # Time the parsing process
                                parsing_start = time.time()
                                parsed_data = parse_analysis(analysis, resume_text, job_description)
                                parsing_time = time.time() - parsing_start
                                st.session_state.total_parsing_time += parsing_time
                                parsing_timer_container.metric("⏱️ Parsing", f"{parsing_time:.2f} seconds")
                                
                                if parsed_data:
                                    # Store the resume text for possible phone number extraction
                                    parsed_data['resume_text'] = resume_text
                                    
                                    # Store the analyzed data in session state for dashboard
                                    st.session_state.results_data.append(parsed_data)
                                    
                                    # Calculate and display time metrics for this resume
                                    resume_time = time.time() - resume_start_time
                                    st.session_state.total_processing_time += resume_time
                                    st.session_state.processed_count += 1
                                    
                                    # Update timer metrics in sidebar
                                    current_timer_container.metric("⏱️ Current Resume", f"{resume_time:.2f} seconds")
                                    
                                    avg_time = st.session_state.total_processing_time / st.session_state.processed_count
                                    avg_timer_container.metric("⏱️ Average Time", f"{avg_time:.2f} seconds/resume")
                                    total_timer_container.metric("⏱️ Total Time", f"{st.session_state.total_processing_time:.2f} seconds")
                                    
                                    # Success message with timing information
                                    st.success(f"Successfully analyzed {uploaded_file.name} in {resume_time:.2f} seconds")
                                    
                                    # Add an expander to show the skill match reasoning
                                    with st.expander("View Skill Matching Details", expanded=False):
                                        st.markdown("### Strong Matches")
                                        st.markdown(f"**Score: {parsed_data['Strong Matches Score']}**")
                                        st.markdown(parsed_data["Strong Matches Reasoning"])
                                        
                                        st.markdown("### Partial Matches")
                                        st.markdown(f"**Score: {parsed_data['Partial Matches Score']}**")
                                        st.markdown(parsed_data["Partial Matches Reasoning"])
                                else:
                                    st.warning(f"Could not extract structured data for {uploaded_file.name}")
                        else:
                            st.error(f"Could not extract text from {uploaded_file.name}")
                            
                    progress_bar.progress((i + 1) / total_files)
                
                # Calculate and show total batch processing time
                batch_time = time.time() - batch_start_time
                progress_bar.progress(1.0)
                
                # Create detailed timing summary
                if st.session_state.processed_count > 0:
                    st.subheader("⏱️ Timing Summary")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Total Batch Time", f"{batch_time:.2f} seconds")
                    with col2:
                        avg_time = st.session_state.total_processing_time / st.session_state.processed_count
                        st.metric("Average Resume Time", f"{avg_time:.2f} seconds")
                    with col3:
                        if st.session_state.total_api_time > 0:
                            api_percentage = (st.session_state.total_api_time / st.session_state.total_processing_time) * 100
                            st.metric("API Call Percentage", f"{api_percentage:.1f}%")
                    
                    st.info(f"✅ Batch processing complete! Total time: {batch_time:.2f} seconds for {len(uploaded_files)} resumes")
                    
                    # Add detailed timing breakdown
                    with st.expander("See detailed timing breakdown", expanded=False):
                        timing_df = pd.DataFrame({
                            "Component": ["API Calls", "PDF Extraction", "Parsing", "Other"],
                            "Total Time (sec)": [
                                round(st.session_state.total_api_time, 2),
                                round(st.session_state.total_extraction_time, 2),
                                round(st.session_state.total_parsing_time, 2),
                                round(st.session_state.total_processing_time - st.session_state.total_api_time - 
                                      st.session_state.total_parsing_time - st.session_state.total_extraction_time, 2)
                            ]
                        })
                        timing_df["Percentage"] = (timing_df["Total Time (sec)"] / st.session_state.total_processing_time * 100).round(1).astype(str) + '%'
                        st.table(timing_df)
        
        # Modify this part to use session state data
        if st.session_state.results_data:
            st.subheader("Analysis Results")
            
            try:
                # Time the dataframe and results creation
                results_start_time = time.time()
                
                # Create DataFrame with the extracted data
                df = pd.DataFrame(st.session_state.results_data)
                
                # Define the key columns for display in the UI
                display_columns = [
                    "Candidate Name", "Total Experience (Years)", "Strong Matches Score", 
                    "Partial Matches Score", "Relevancy Score (0-100)", "Overall Weighted Score",
                    "College Rating", "Job Stability", "Latest Company",
                    "Leadership Skills", "International Team Experience",
                    "Competitor Experience", "Selection Recommendation"
                ]
                
                # Show all columns that exist in our dataframe
                available_columns = [col for col in display_columns if col in df.columns]
                
                if available_columns:
                    st.dataframe(df[available_columns])
                else:
                    st.warning("No columns to display. Please check the AI response format.")
                    st.write("DataFrame columns:", df.columns.tolist())
                
                results_time = time.time() - results_start_time
                
                # For the Excel export, we want all columns
                export_columns = [
                    "Candidate Name", "Total Experience (Years)", 
                    "Strong Matches Score", "Strong Matches Reasoning", 
                    "Partial Matches Score", "Partial Matches Reasoning", 
                    "Relevancy Score (0-100)", "All Tech Skills", "Relevant Tech Skills",
                    "Degree", "College/University", "Job Applying For", "College Rating", 
                    "Job Stability", "Latest Company", "Leadership Skills", 
                    "International Team Experience", "Notice Period", "LinkedIn URL", 
                    "Portfolio URL", "Work History", "Competitor Experience",
                    "Overall Weighted Score", "Selection Recommendation"
                ]
                
                # Available export columns
                available_export_columns = [col for col in export_columns if col in df.columns]
                
                with st.spinner("Preparing Excel file..."):
                    excel_start_time = time.time()
                    try:
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmpfile:
                            # Make sure we're saving the dataframe with all available columns
                            if available_export_columns:
                                export_df = df[available_export_columns]
                            else:
                                export_df = df  # Use all columns if our expected ones aren't found
                            
                            export_df.to_excel(tmpfile.name, index=False, sheet_name='Resume Analysis', engine='openpyxl')
                            wb = openpyxl.load_workbook(tmpfile.name)
                            
                            if available_export_columns:
                                wb = format_excel_workbook(wb, available_export_columns)
                            else:
                                # If no expected columns, use whatever columns are in the dataframe
                                wb = format_excel_workbook(wb, df.columns.tolist())
                            
                            wb.save(tmpfile.name)
                            tmpfile_path = tmpfile.name
                        
                        excel_time = time.time() - excel_start_time
                        st.success(f"Excel report ready! (Prepared in {excel_time:.2f} seconds)")
                        
                        with open(tmpfile_path, "rb") as file:
                            file_data = file.read()
                            st.download_button(
                                label="📥 Download Complete Resume Analysis Report",
                                data=file_data,
                                file_name=f"resume_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                        os.unlink(tmpfile_path)
                    except Exception as e:
                        st.error(f"Error creating Excel file: {str(e)}")
                        st.info("You can still see the results in the table above.")
            except Exception as e:
                st.error(f"Error processing results: {str(e)}")
                import traceback
                st.error(traceback.format_exc())
    except Exception as e:
        st.error(f"An unexpected error occurred: {str(e)}")
        import traceback
        st.error(traceback.format_exc())

if __name__ == "__main__":
    main()# Dashboard tab functionality
def dashboard_tab():
    # Check if we have results data in session state
    if 'results_data' not in st.session_state or not st.session_state.results_data:
        st.info("No candidates have been analyzed yet. Please analyze resumes in the Resume Analysis tab first.")
        return
    
    # Get the data from session state
    results_data = st.session_state.results_data
    
    # Extract phone numbers if not already present
    for candidate in results_data:
        if 'Phone Number' not in candidate or not candidate['Phone Number']:
            # Try to extract from resume text if available
            if 'resume_text' in candidate:
                candidate['Phone Number'] = extract_phone_number(candidate['resume_text'])
            else:
                candidate['Phone Number'] = "Not Available"
    
    # Create DataFrame with the key columns for the dashboard
    df = pd.DataFrame(results_data)
    
    # Select columns for dashboard
    dashboard_columns = ["Candidate Name", "Phone Number", "Job Applying For", 
                        "Total Experience (Years)", "Overall Weighted Score", 
                        "Selection Recommendation"]
    
    # Filter to only available columns
    available_columns = [col for col in dashboard_columns if col in df.columns]
    
    if len(available_columns) < 3:  # Not enough data to display
        st.warning("Insufficient data for dashboard. Please ensure the analysis includes candidate names and scores.")
        return
    
    # Sort by Overall Weighted Score in descending order
    if "Overall Weighted Score" in df.columns:
        df["Overall Weighted Score"] = pd.to_numeric(df["Overall Weighted Score"], errors='coerce')
        df = df.sort_values(by="Overall Weighted Score", ascending=False)
    
    # Add styling to the dataframe
    def highlight_recommendation(val):
        if isinstance(val, str):
            if "Strong Fit" in val:
                return 'background-color: #C6EFCE; color: #006100'
            elif "Good Fit" in val:
                return 'background-color: #C6EFCE; color: #006100'
            elif "Consider" in val:
                return 'background-color: #FFEB9C; color: #9C5700'
            elif "Weak Fit" in val:
                return 'background-color: #FFCC00; color: #9C5700'
            elif "Reject" in val:
                return 'background-color: #FFC7CE; color: #9C0006'
        return ''
    
    # Apply the styling
    styled_df = df[available_columns].style.applymap(
        highlight_recommendation, 
        subset=['Selection Recommendation'] if 'Selection Recommendation' in available_columns else []
    )
    
    # Dashboard metrics at the top
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Total Candidates", len(df))
    
    with col2:
        if "Selection Recommendation" in df.columns:
            recommended_count = df[df["Selection Recommendation"].str.contains("Strong Fit|Good Fit", na=False)].shape[0]
            st.metric("Recommended Candidates", recommended_count)
    
    with col3:
        if "Overall Weighted Score" in df.columns:
            avg_score = df["Overall Weighted Score"].mean()
            st.metric("Average Score", f"{avg_score:.1f}")
    
    with col4:
        if "Total Experience (Years)" in df.columns:
            df["Total Experience (Years)"] = pd.to_numeric(df["Total Experience (Years)"], errors='coerce')
            avg_exp = df["Total Experience (Years)"].mean()
            st.metric("Average Experience", f"{avg_exp:.1f} years")
    
    # Add visualization - Score distribution
    st.subheader("Score Distribution")
    
    if "Overall Weighted Score" in df.columns:
        # Create score bins
        score_bins = [0, 40, 55, 70, 85, 100]
        score_labels = ['Reject', 'Weak Fit', 'Consider', 'Good Fit', 'Strong Fit']
        
        df['Score Category'] = pd.cut(df["Overall Weighted Score"], bins=score_bins, labels=score_labels, right=False)
        
        # Count candidates in each category
        category_counts = df['Score Category'].value_counts().reset_index()
        category_counts.columns = ['Category', 'Count']
        
        # Sort by the order in score_labels
        category_counts['Category'] = pd.Categorical(category_counts['Category'], categories=score_labels, ordered=True)
        category_counts = category_counts.sort_values('Category')
        
        # Create the chart
        fig = px.bar(category_counts, x='Category', y='Count', 
                     color='Category',
                     color_discrete_map={
                         'Strong Fit': '#4CAF50',
                         'Good Fit': '#8BC34A',
                         'Consider': '#FFEB3B',
                         'Weak Fit': '#FF9800',
                         'Reject': '#F44336'
                     },
                     text='Count')
        
        fig.update_layout(height=400, width=700)
        st.plotly_chart(fig, use_container_width=True)
    
    # Main candidates table
    st.subheader("Candidates Overview")
    st.dataframe(styled_df, height=400, use_container_width=True)
    
    # Add export functionality
    if not df.empty:
        csv = df[available_columns].to_csv(index=False)
        st.download_button(
            label="📥 Download Candidates Overview",
            data=csv,
            file_name=f"candidates_overview_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv"
        )# Format Excel with styling and organization
def format_excel_workbook(wb, columns):
    try:
        ws = wb.active
        
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        normal_font = Font(name='Calibri', size=11)
        normal_alignment = Alignment(vertical='center', wrap_text=True)
        
        score_alignment = Alignment(horizontal='center', vertical='center')
        
        url_font = Font(name='Calibri', size=11, color='0000FF', underline='single')
        
        competitor_yes_font = Font(name='Calibri', size=11, bold=True, color='FF0000')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply formatting to headers
        for col_num, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Apply formatting to data cells
        max_row = ws.max_row if ws.max_row else 1  # Protect against empty worksheet
        max_col = ws.max_column if ws.max_column else 1
        
        for row in range(2, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if not cell.value:  # Skip empty cells
                    continue
                    
                cell.font = normal_font
                cell.alignment = normal_alignment
                cell.border = thin_border
                
                if col <= len(columns):  # Ensure we don't go out of bounds
                    column_name = columns[col-1]
                    
                    if any(term in column_name for term in ["Score", "Recommendation", "Job Stability"]):
                        cell.alignment = score_alignment
                        
                        if cell.value not in ["Not Available", None, ""]:
                            try:
                                if any(term in column_name for term in ["Score", "Job Stability"]):
                                    score_value = float(cell.value)
                                    if score_value >= 75 or (column_name == "Job Stability" and score_value >= 8):
                                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green
                                    elif score_value >= 50 or (column_name == "Job Stability" and score_value >= 6):
                                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Yellow
                                    else:
                                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Red
                            except (ValueError, TypeError):
                                pass
                    
                    if column_name == "College Rating" and cell.value not in ["Not Available", None, ""]:
                        if "premium" in str(cell.value).lower() and "non" not in str(cell.value).lower():
                            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green
                        elif "non-premium" in str(cell.value).lower():
                            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Yellow
                    
                    if column_name == "Selection Recommendation" and cell.value not in ["Not Available", None, ""]:
                        if "Strong Fit" in str(cell.value) or "Good Fit" in str(cell.value):
                            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green
                        elif "Consider" in str(cell.value):
                            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Yellow
                        elif "Weak Fit" in str(cell.value):
                            cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')  # Orange
                        elif "Reject" in str(cell.value):
                            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Red
                    
                    if column_name in ["LinkedIn URL", "Portfolio URL"] and cell.value not in ["Not Available", None, ""]:
                        cell.font = url_font
                        try:
                            cell.hyperlink = cell.value
                        except Exception:
                            # Fallback if hyperlink fails
                            pass
                    
                    if column_name == "Competitor Experience" and cell.value not in ["Not Available", None, ""]:
                        if str(cell.value).lower().startswith("yes"):
                            cell.font = competitor_yes_font
                            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Red background
        
        # Set column widths
        for col_num, column in enumerate(columns, 1):
            column_letter = openpyxl.utils.get_column_letter(col_num)
            if any(term in column for term in ["Skills", "Reasoning", "Leadership", "International", "Experience", "Work History"]):
                ws.column_dimensions[column_letter].width = 40
            elif any(term in column for term in ["Recommendation", "Notice", "Company", "College", "URL"]):
                ws.column_dimensions[column_letter].width = 30
            else:
                ws.column_dimensions[column_letter].width = 18
        
        # Freeze the top row
        ws.freeze_panes = "A2"
        
        return wb
    except Exception as e:
        st.error(f"Error formatting Excel: {str(e)}")
        # Return the unformatted workbook as fallback
        return wbimport streamlit as st
from groq import Groq
import PyPDF2
import os
import re
import pandas as pd
from dotenv import load_dotenv
from datetime import datetime
import tempfile
import openpyxl
from io import BytesIO
import time  # For timing functionality
import plotly.express as px

# Initialize AI client
def initialize_groq_client():
    try:
        return Groq(api_key=os.environ.get("GROQ_API_KEY"))
    except Exception as e:
        st.error(f"Failed to initialize Groq client: {str(e)}")
        return None

# Extract text from PDF
def extract_text_from_pdf(pdf_file):
    try:
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        text = "\n".join([page.extract_text() for page in pdf_reader.pages if page.extract_text()])
        return text if text else None
    except Exception as e:
        st.error(f"Error extracting text from PDF: {str(e)}")
        return None

# Define Planful competitors
def get_planful_competitors():
    return [
        "Anaplan", "Workday Adaptive Planning", "Oracle EPM", "Oracle Hyperion", 
        "SAP BPC", "IBM Planning Analytics", "TM1", "Prophix", "Vena Solutions", 
        "Jedox", "OneStream", "Board", "Centage", "Solver", "Kepion", "Host Analytics",
        "CCH Tagetik", "Infor CPM", "Syntellis", "Longview"
    ]

# Extract LinkedIn URL directly from resume text
def extract_linkedin_url(text):
    if not text:
        return ""
    
    patterns = [
        r'https?://(?:www\.)?linkedin\.com/in/[\w-]+(?:/[\w-]+)*',
        r'linkedin\.com/in/[\w-]+(?:/[\w-]+)*',
        r'www\.linkedin\.com/in/[\w-]+(?:/[\w-]+)*',
        r'linkedin:\s*https?://(?:www\.)?linkedin\.com/in/[\w-]+',
    ]
    
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        if matches:
            url = matches[0]
            if not url.startswith('http'):
                url = 'https://' + ('' if url.startswith('www.') or url.startswith('linkedin.com') else 'www.') + url
                if url.startswith('https://linkedin.com'):
                    url = url.replace('https://linkedin.com', 'https://www.linkedin.com')
            url = re.sub(r'[.,;:)\s]+$', '', url)
            return url
    
    linkedin_mention = re.search(r'linkedin[\s:]*([^\s]+)', text, re.IGNORECASE)
    if linkedin_mention:
        potential_url = linkedin_mention.group(1)
        if '.' in potential_url and '/' in potential_url:
            url = re.sub(r'[.,;:)\s]+$', '', potential_url)
            if not url.startswith('http'):
                url = 'https://' + ('' if url.startswith('www.') else 'www.') + url
            return url
    
    return ""

# Extract phone number from resume text
def extract_phone_number(text):
    if not text:
        return "Not Available"
    
    # Common phone number patterns
    patterns = [
        r'\b(?:\+\d{1,3}[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}\b',  # (123) 456-7890, 123-456-7890
        r'\b\d{10}\b',  # 1234567890
        r'\b\d{3}[-.\s]\d{3}[-.\s]\d{4}\b',  # 123-456-7890, 123.456.7890
        r'\b\+\d{1,3}\s?\d{6,14}\b'  # International format: +1 1234567890
    ]
    
    for pattern in patterns:
        matches = re.findall(pattern, text)
        if matches:
            return matches[0]
    
    return "Not Available"

# Calculate the individual scores and overall score based on the improved algorithm
def calculate_scores(parsed_data, required_experience=3, stability_threshold=2):
    try:
        scores = {}
        
        # Strong Matches Score - direct from the AI analysis of exact skill matches
        strong_matches_val = parsed_data.get("Strong Matches Score", "0")
        try:
            scores["strong_matches"] = float(strong_matches_val)
        except (ValueError, TypeError):
            scores["strong_matches"] = 0
            
        # Partial Matches Score - direct from the AI analysis of related skills
        partial_matches_val = parsed_data.get("Partial Matches Score", "0")
        try:
            scores["partial_matches"] = float(partial_matches_val)
        except (ValueError, TypeError):
            scores["partial_matches"] = 0
            
        # Calculate relevancy score as a weighted sum of strong and partial matches
        # Give more weight to strong matches (70%) than partial matches (30%)
        weighted_strong = scores["strong_matches"] * 0.7
        weighted_partial = scores["partial_matches"] * 0.3
        
        # Final relevancy score is the sum of weighted strong and partial matches
        scores["relevancy"] = min(weighted_strong + weighted_partial, 100)
        
        # Update the parsed data with our calculated relevancy score
        parsed_data["Relevancy Score (0-100)"] = str(round(scores["relevancy"], 1))
        
        # Experience calculation - based on required years
        experience_val = parsed_data.get("Total Experience (Years)", "0")
        try:
            candidate_exp = float(experience_val)
            # More nuanced experience score:
            # - Below required: proportional score up to 70%
            # - At required: 80%
            # - Above required: bonus points up to 100%
            if candidate_exp < required_experience:
                scores["experience"] = min((candidate_exp / required_experience) * 70, 70)
            elif candidate_exp == required_experience:
                scores["experience"] = 80
            else:
                # Additional experience gives bonus points, with diminishing returns
                bonus = min(((candidate_exp - required_experience) / 2) * 20, 20)
                scores["experience"] = 80 + bonus
        except (ValueError, TypeError):
            scores["experience"] = 0
            
        # Job stability - how long candidates typically stay at jobs
        stability_val = parsed_data.get("Job Stability", "0")
        try:
            job_stability = float(stability_val)
            if job_stability <= 10:  # If rated on 1-10 scale
                scores["stability"] = job_stability * 10  # Convert to 100-point scale
            else:  # If provided as average years
                # Convert years to score: 
                # - Less than 1 year: proportional score up to 50
                # - 1-2 years: 50-85
                # - 2+ years: 85-100
                if job_stability < 1:
                    scores["stability"] = (job_stability * 50)
                elif job_stability < 2:
                    scores["stability"] = 50 + ((job_stability - 1) * 35)
                else:
                    scores["stability"] = 85 + min(((job_stability - 2) * 7.5), 15)
        except (ValueError, TypeError):
            scores["stability"] = 0
            
        # College rating score
        college_rating = parsed_data.get("College Rating", "")
        if college_rating:
            if "premium" in college_rating.lower() and "non" not in college_rating.lower():
                scores["college"] = 100
            elif "non-premium" in college_rating.lower():
                scores["college"] = 70
            else:
                scores["college"] = 40
        else:
            scores["college"] = 20
            
        # Leadership score - based on presence of leadership experience
        leadership_skills = parsed_data.get("Leadership Skills", "")
        if leadership_skills:
            leadership_keywords = ["led", "managed", "directed", "leadership", "head", "team lead", 
                                "supervisor", "manager", "chief", "director", "lead"]
            
            if any(word in leadership_skills.lower() for word in leadership_keywords):
                scores["leadership"] = 100
            else:
                # Check for partial leadership indicators
                partial_leadership = ["coordinated", "facilitated", "organized", "spearheaded", "guided"]
                if any(word in leadership_skills.lower() for word in partial_leadership):
                    scores["leadership"] = 50
                else:
                    scores["leadership"] = 0
        else:
            scores["leadership"] = 0
            
        # International experience score
        international_exp = parsed_data.get("International Team Experience", "")
        if international_exp:
            international_keywords = ["yes", "international", "global", "worldwide", "multinational", 
                                    "cross-border", "overseas", "remote teams", "offshore"]
            
            if any(word in international_exp.lower() for word in international_keywords):
                # Look for deeper international experience
                deep_int_exp = ["led international", "managed global", "cross-cultural", "multiple countries"]
                if any(phrase in international_exp.lower() for phrase in deep_int_exp):
                    scores["international"] = 100
                else:
                    scores["international"] = 80
            else:
                scores["international"] = 0
        else:
            scores["international"] = 0
            
        # Competitor experience score - more nuanced based on specific competitors
        competitor_exp = parsed_data.get("Competitor Experience", "")
        if competitor_exp and competitor_exp.lower().startswith("yes"):
            # Premium competitors get higher scores
            premium_competitors = ["anaplan", "workday", "oracle", "sap", "onestream"]
            if any(comp in competitor_exp.lower() for comp in premium_competitors):
                scores["competitor"] = 100
            else:
                scores["competitor"] = 70
        else:
            scores["competitor"] = 0
            
        # Calculate weighted overall score with adjusted weights
        overall_score = (
            (0.40 * scores["relevancy"]) +        # Skills relevancy is most important
            (0.15 * scores["experience"]) +       # Years of experience
            (0.12 * scores["stability"]) +        # Job stability slightly more important
            (0.10 * scores["college"]) +          # Education background
            (0.10 * scores["leadership"]) +       # Leadership abilities
            (0.08 * scores["international"]) +    # International experience slightly less weight
            (0.05 * scores["competitor"])         # Competitor experience
        )
        
        # Enhanced recommendation categories
        if overall_score >= 85:
            recommendation = "Strong Fit ✅ - Priority interview"
        elif overall_score >= 70:
            recommendation = "Good Fit ✅ - Recommend interview"
        elif overall_score >= 55:
            recommendation = "Consider 🤔 - Further screening needed"
        elif overall_score >= 40:
            recommendation = "Weak Fit ⚠️ - Only interview if candidate pool is limited"
        else:
            recommendation = "Reject ❌ - Does not meet minimum criteria"
            
        return overall_score, recommendation, scores
    
    except Exception as e:
        st.error(f"Error calculating scores: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return 0, "Error in calculation", {}

# Analyze resume with detailed skill matching
def analyze_resume(client, resume_text, job_description):
    if not client:
        return None
        
    competitors = get_planful_competitors()
    competitors_list = ", ".join(competitors)
    
    # This prompt is focused on detailed skill extraction and matching
    prompt = f"""
    You are an experienced HR Consultant analyzing a candidate resume against a job description for a technical role.
    Your task is to carefully identify skills and match them between the job description and resume.

    First, extract a comprehensive list of ALL required skills, qualifications, and technologies from the job description.
    Then thoroughly analyze the resume to identify skills that exactly match or are related to the job requirements.

    Provide your analysis in the following format:

    Candidate Name: [Full name from resume]
    Total Experience (Years): [Total years of professional experience]

    Strong Matches Score (0-100): [IMPORTANT: Assign a numeric score based on exact skill matches]
    Strong Matches Reasoning: [List each exact skill match with evidence from resume]

    Partial Matches Score (0-100): [IMPORTANT: Assign a numeric score based on related/transferable skills]
    Partial Matches Reasoning: [List each related skill with explanation]

    Relevancy Score (0-100): [Calculate as: 70% of Strong Matches + 30% of Partial Matches]

    All Tech Skills: [All technical skills mentioned in resume]
    Relevant Tech Skills: [Only skills relevant to this job]
    Degree: [Highest degree earned]
    College/University: [Institution name]
    Job Applying For: [Job title/ID from description]
    College Rating: [Rate as "Premium" or "Non-Premium"]
    Job Stability: [Rate 1-10 based on average tenure]
    Latest Company: [Most recent employer]
    Leadership Skills: [Leadership experience details]
    International Team Experience: [Details about global team experience]
    Notice Period: [When candidate can join]
    LinkedIn URL: [LinkedIn profile if mentioned]
    Portfolio URL: [Portfolio/GitHub if mentioned]
    Work History: [Summary of previous roles]
    Competitor Experience: [Only "Yes - [Company]" if worked at: {competitors_list}. Otherwise leave blank]

    SCORING INSTRUCTIONS:
    - For Strong Matches Score: Count the number of exact skill matches, divide by total required skills, multiply by 100
    - For Partial Matches Score: Count related/transferable skills, evaluate relevance (50-80% per skill), average them
    - A score of 0 should ONLY be given if absolutely NO matches are found
    - Be generous with partial matches - if a skill is conceptually related, count it
    - Do not artificially deflate scores - real-world recruitment values transferable skills

    Resume:
    {resume_text}

    Job Description:
    {job_description}
    """
    
    try:
        # Track time for API call
        api_call_start = time.time()
        
        response = client.chat.completions.create(
            model="mixtral-8x7b-32768",
            messages=[
                {"role": "system", "content": "You are an expert HR consultant with years of technical recruitment experience. Your specialty is identifying transferable skills between different technologies and roles."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,  # Lower temperature for consistent analysis
            max_tokens=3500   # Increased to allow for detailed analysis
        )
        
        # Calculate API call time
        api_call_time = time.time() - api_call_start
        
        ai_response = response.choices[0].message.content
        
        # Add API call time to the response for later use
        ai_response = f"API call time: {api_call_time:.2f} seconds\n\n" + ai_response
        
        # Debug info to see raw output for troubleshooting
        with st.expander("AI Analysis (Debug)", expanded=False):
            st.write(f"API call time: {api_call_time:.2f} seconds")
            st.write(ai_response[:500] + "..." if len(ai_response) > 500 else ai_response)
            
            # Check for score mentions in the response
            strong_score_match = re.search(r'Strong Matches Score:?\s*(\d+)', ai_response)
            partial_score_match = re.search(r'Partial Matches Score:?\s*(\d+)', ai_response)
            
            if strong_score_match:
                st.write(f"✅ Strong Matches Score detected: {strong_score_match.group(1)}")
            else:
                st.write("❌ Strong Matches Score not found in response")
                
            if partial_score_match:
                st.write(f"✅ Partial Matches Score detected: {partial_score_match.group(1)}")
            else:
                st.write("❌ Partial Matches Score not found in response")
        
        return ai_response
    except Exception as e:
        st.error(f"Error during analysis: {str(e)}")
        return None

# Clean text by removing formatting
def clean_text(text):
    if not text or text == "Not Available":
        return text
        
    # Remove markdown formatting
    text = re.sub(r'\*\*(.*?)\*\*', r'\1', text)  # Bold
    text = re.sub(r'\*(.*?)\*', r'\1', text)      # Italic
    text = re.sub(r'__(.*?)__', r'\1', text)      # Underline
    text = re.sub(r'_(.*?)_', r'\1', text)        # Italic alternative
    text = re.sub(r'`(.*?)`', r'\1', text)        # Code
    
    # Remove bullet points and numbering
    text = re.sub(r'^\s*[-•*]\s+', '', text, flags=re.MULTILINE)
    text = re.sub(r'^\s*\d+\.\s+', '', text, flags=re.MULTILINE)
    
    return text.strip()

# Check for competitor mentions in work history
def check_competitor_experience(work_history, competitor_list):
    if not work_history or work_history == "Not Available":
        return ""
    
    for competitor in competitor_list:
        # Case-insensitive word boundary match
        pattern = r'\b' + re.escape(competitor.lower()) + r'\b'
        if re.search(pattern, work_history.lower()):
            return f"Yes - {competitor}"
    
    return ""

# Improved calculate_skills_scores function with detailed reasoning
def calculate_skills_scores(resume_text, job_description):
    """
    Provides detailed reasoning for skill matches and scores when the AI fallback is used.
    Returns strong score, partial score, and detailed reasoning for both.
    """
    # Normalize text for comparison
    resume_lower = resume_text.lower()
    jd_lower = job_description.lower()
    
    # List of common technical skills to check for
    common_skills = [
        "python", "java", "javascript", "c++", "c#", ".net", "php", "ruby", "swift",
        "sql", "mysql", "postgresql", "mongodb", "oracle", "database", 
        "aws", "azure", "gcp", "cloud", "docker", "kubernetes", "devops",
        "html", "css", "react", "angular", "vue", "node.js", "django",
        "machine learning", "ai", "data science", "tensorflow", "pytorch",
        "excel", "powerbi", "tableau", "power bi", "data visualization",
        "agile", "scrum", "jira", "project management", "pmp",
        "linux", "unix", "windows", "git", "github", "gitlab",
        "api", "rest", "graphql", "microservices", "serverless",
        # Common business and finance skills
        "financial analysis", "budgeting", "forecasting", "accounting",
        "strategic planning", "business development", "marketing", "sales",
        "customer relationship management", "crm", "sap", "erp",
        # Common soft skills
        "communication", "leadership", "teamwork", "problem solving",
        "critical thinking", "time management", "organization"
    ]
    
    # Define related skills (skills that are similar or related to each other)
    related_skills = {
        "python": ["django", "flask", "pandas", "numpy", "data science", "machine learning", "AI"],
        "java": ["spring", "hibernate", "j2ee", "android"],
        "javascript": ["typescript", "node.js", "react", "angular", "vue", "front-end"],
        "sql": ["mysql", "postgresql", "oracle", "database", "data analysis"],
        "aws": ["cloud", "azure", "gcp", "devops", "infrastructure"],
        "docker": ["kubernetes", "containers", "devops", "microservices"],
        "tableau": ["power bi", "data visualization", "analytics", "reporting"],
        "excel": ["spreadsheets", "data analysis", "financial modeling"],
        "agile": ["scrum", "kanban", "jira", "project management"],
        "machine learning": ["ai", "data science", "deep learning", "nlp"],
    }
    
    # Find all skills mentioned in the job description
    jd_skills = []
    for skill in common_skills:
        # Use word boundaries to ensure we're matching whole words
        pattern = r'\b' + re.escape(skill) + r'\b'
        if re.search(pattern, jd_lower):
            jd_skills.append(skill)
    
    # Find exact matches in the resume
    exact_matches = []
    for skill in jd_skills:
        pattern = r'\b' + re.escape(skill) + r'\b'
        if re.search(pattern, resume_lower):
            exact_matches.append(skill)
    
    # Find related matches
    related_matches = []
    for jd_skill in jd_skills:
        if jd_skill in exact_matches:
            continue  # Skip if already an exact match
            
        # Check if any related skills are in the resume
        if jd_skill in related_skills:
            for related_skill in related_skills[jd_skill]:
                pattern = r'\b' + re.escape(related_skill.lower()) + r'\b'
                if re.search(pattern, resume_lower) and related_skill.lower() not in [match.lower() for match in exact_matches]:
                    related_matches.append(f"{related_skill} (related to {jd_skill})")
    
    # Additional resume skills that might be transferable
    additional_resume_skills = []
    for skill in common_skills:
        if skill not in jd_skills:  # Don't include skills already counted
            pattern = r'\b' + re.escape(skill) + r'\b'
            if re.search(pattern, resume_lower):
                for jd_skill in jd_skills:
                    if skill in related_skills.get(jd_skill, []) or jd_skill in related_skills.get(skill, []):
                        related_match = f"{skill} (transferable to {jd_skill})"
                        if related_match not in related_matches:
                            related_matches.append(related_match)
    
    # Calculate scores
    if not jd_skills:  # No skills found in JD
        strong_score = 50  # Default middle score
        strong_reasoning = "No specific technical skills identified in the job description. Using default middle score."
    else:
        strong_score = (len(exact_matches) / len(jd_skills)) * 100
        strong_reasoning = f"Found {len(exact_matches)} exact matches out of {len(jd_skills)} required skills ({strong_score:.1f}%).\n\n"
        strong_reasoning += "Exact skill matches:\n"
        
        if exact_matches:
            for skill in exact_matches:
                strong_reasoning += f"- {skill.upper()}: Found in both resume and job description\n"
        else:
            strong_reasoning += "- No exact skill matches found\n"
        
        strong_reasoning += f"\nMissing skills from job description:\n"
        for skill in jd_skills:
            if skill not in exact_matches:
                strong_reasoning += f"- {skill}\n"
    
    # Related skills score - calculate based on the number of related matches
    max_related_score = min(80, strong_score + 20)  # Cap at 80 or 20% higher than strong
    min_related_score = max(30, strong_score * 0.6)  # Floor at 30 or 60% of strong
    
    if related_matches:
        # Scale the related score based on the number of related matches relative to missing skills
        missing_skills_count = len(jd_skills) - len(exact_matches)
        if missing_skills_count > 0:
            related_coverage = min(1.0, len(related_matches) / missing_skills_count)
            partial_score = min_related_score + (max_related_score - min_related_score) * related_coverage
        else:
            partial_score = max_related_score
    else:
        partial_score = min_related_score
    
    partial_reasoning = f"Found {len(related_matches)} related/transferable skills ({partial_score:.1f}%).\n\n"
    partial_reasoning += "Related skill matches:\n"
    
    if related_matches:
        for skill in related_matches:
            partial_reasoning += f"- {skill}\n"
    else:
        partial_reasoning += "- No related skill matches found\n"
    
    return round(strong_score), round(partial_score), strong_reasoning, partial_reasoning

# Parse AI response with improved extraction logic
def parse_analysis(analysis, resume_text=None, job_description=None):
    try:
        if not analysis:
            return None
            
        # Definition of expected fields with exact matches and alternative formats
        expected_fields = {
            "Candidate Name": ["candidate name", "candidate's name", "name"],
            "Total Experience (Years)": ["total experience (years)", "total experience", "experience (years)", "years of experience"],
            "Relevancy Score (0-100)": ["relevancy score (0-100)", "relevancy score", "relevance score"],
            "Strong Matches Score": ["strong matches score", "strong match score", "strong matches"],
            "Strong Matches Reasoning": ["strong matches reasoning", "strong match reasoning"],
            "Partial Matches Score": ["partial matches score", "partial match score", "partial matches"],
            "Partial Matches Reasoning": ["partial matches reasoning", "partial match reasoning"],
            "All Tech Skills": ["all tech skills", "all technical skills"],
            "Relevant Tech Skills": ["relevant tech skills", "relevant technical skills"],
            "Degree": ["degree", "highest degree", "qualification"],
            "College/University": ["college/university", "university", "college", "institution"],
            "Job Applying For": ["job applying for", "job id", "position applying for", "role applying for"],
            "College Rating": ["college rating", "university rating", "institution rating"],
            "Job Stability": ["job stability", "employment stability"],
            "Latest Company": ["latest company", "current company", "most recent company"],
            "Leadership Skills": ["leadership skills", "leadership experience", "leadership"],
            "International Team Experience": ["international team experience", "global team experience", "international experience"],
            "Notice Period": ["notice period", "joining availability", "availability to join"],
            "LinkedIn URL": ["linkedin url", "linkedin profile", "linkedin", "linkedin link"],
            "Portfolio URL": ["portfolio url", "portfolio", "github url", "github", "personal website", "personal url", "website"],
            "Work History": ["work history", "employment history", "companies worked for", "previous companies"],
            "Competitor Experience": ["competitor experience", "worked for competitor", "competitor", "competition experience"],
        }
        
        # Create a dictionary to store the extracted values
        result = {field: "Not Available" for field in expected_fields}
        
        # Split the AI output into lines for processing
        lines = analysis.split('\n')
        
        # First pass: direct pattern matching for scores
        # This has higher priority because we want to ensure we catch these values
        strong_match = re.search(r'Strong Matches Score:?\s*(\d+(?:\.\d+)?)', analysis)
        if strong_match:
            result["Strong Matches Score"] = strong_match.group(1)
            
        partial_match = re.search(r'Partial Matches Score:?\s*(\d+(?:\.\d+)?)', analysis)
        if partial_match:
            result["Partial Matches Score"] = partial_match.group(1)
        
        # Second pass: structured field extraction
        current_field = None
        current_value = []
        
        for i, line in enumerate(lines):
            line = line.strip()
            if not line:  # Skip empty lines
                continue
                
            # Check if this line starts a new field
            new_field_found = False
            
            if ':' in line:
                parts = line.split(':', 1)
                key = parts[0].strip().lower()
                value = parts[1].strip()
                
                # Check if this matches any of our expected fields
                for field, alternatives in expected_fields.items():
                    if key in alternatives:
                        # If we were building a previous field value, save it
                        if current_field and current_value:
                            result[current_field] = '\n'.join(current_value)
                            
                        # Start the new field
                        current_field = field
                        current_value = [value] if value else []
                        new_field_found = True
                        break
            
            # If this line doesn't start a new field and we're in the middle of a field, append to current value
            if not new_field_found and current_field and line:
                # Only append if the line doesn't look like it might be a mislabeled field
                if ':' not in line or line.split(':', 1)[0].strip().lower() not in [alt for alts in expected_fields.values() for alt in alts]:
                    current_value.append(line)
            
            # If we're at the last line and have an active field, save it
            if i == len(lines) - 1 and current_field and current_value:
                result[current_field] = '\n'.join(current_value)
        
        # Third pass: extract numeric values from fields
        numeric_fields = ["Total Experience (Years)", "Relevancy Score (0-100)", "Strong Matches Score", 
                         "Partial Matches Score", "Job Stability"]
        
        for field in numeric_fields:
            if result[field] != "Not Available":
                # Try to extract a numeric value
                matches = re.search(r'(\d+(?:\.\d+)?)', result[field])
                if matches:
                    result[field] = matches.group(1)
        
        # Special handling for Job Stability
        if result["Job Stability"] != "Not Available" and not re.match(r'^\d+(?:\.\d+)?$', result["Job Stability"]):
            # Try to extract a number from the text
            matches = re.search(r'(\d+(?:\.\d+)?)/10', result["Job Stability"])
            if matches:
                result[field] = matches.group(1)
            else:
                matches = re.search(r'(\d+(?:\.\d+)?)', result["Job Stability"])
                if matches:
                    result[field] = matches.group(1)
        
        # IMPORTANT FALLBACK: If we still don't have scores, calculate them manually
        if (result["Strong Matches Score"] == "Not Available" or result["Strong Matches Score"] == "0") and \
           (result["Partial Matches Score"] == "Not Available" or result["Partial Matches Score"] == "0") and \
           resume_text and job_description:
            # Manually calculate scores as fallback with detailed reasoning
            strong_score, partial_score, strong_reasoning, partial_reasoning = calculate_skills_scores(resume_text, job_description)
            result["Strong Matches Score"] = str(strong_score)
            result["Partial Matches Score"] = str(partial_score)
            result["Strong Matches Reasoning"] = strong_reasoning
            result["Partial Matches Reasoning"] = partial_reasoning
        
        # Normalize College Rating
        if result["College Rating"] != "Not Available":
            if "premium" in result["College Rating"].lower():
                result["College Rating"] = "Premium"
            elif "non" in result["College Rating"].lower() or "not" in result["College Rating"].lower():
                result["College Rating"] = "Non-Premium"
        
        # Normalize International Team Experience
        if result["International Team Experience"] != "Not Available":
            if any(word in result["International Team Experience"].lower() for word in ["yes", "has", "worked", "experience"]):
                if len(result["International Team Experience"]) < 5:  # Just "Yes" or similar
                    result["International Team Experience"] = "Yes"
            elif any(word in result["International Team Experience"].lower() for word in ["no", "not", "none"]):
                if len(result["International Team Experience"]) < 5:  # Just "No" or similar
                    result["International Team Experience"] = "No"
        
        # Handle LinkedIn URL extraction
        if resume_text and (result["LinkedIn URL"] == "Not Available" or not result["LinkedIn URL"]):
            result["LinkedIn URL"] = extract_linkedin_url(resume_text)
        elif result["LinkedIn URL"] != "Not Available":
            linkedin_match = re.search(r'https?://(?:www\.)?linkedin\.com/in/[\w-]+(?:/[\w-]+)*', result["LinkedIn URL"])
            if linkedin_match:
                result["LinkedIn URL"] = linkedin_match.group(0)
            else:
                extracted_url = extract_linkedin_url(result["LinkedIn URL"])
                if extracted_url:
                    result["LinkedIn URL"] = extracted_url
        
        # Clean up Portfolio URL
        if result["Portfolio URL"] != "Not Available":
            portfolio_match = re.search(r'https?://(?:www\.)?(?:github\.com|gitlab\.com|bitbucket\.org|behance\.net|dribbble\.com|[\w-]+\.(?:com|io|org|net))/\S+', result["Portfolio URL"])
            if portfolio_match:
                result["Portfolio URL"] = portfolio_match.group(0)
            elif "not available" in result["Portfolio URL"].lower() or "not found" in result["Portfolio URL"].lower() or "not mentioned" in result["Portfolio URL"].lower():
                result["Portfolio URL"] = ""
        else:
            result["Portfolio URL"] = ""
        
        # Use Latest Company if Work History is not available
        if result["Work History"] == "Not Available" and "Latest Company" in result and result["Latest Company"] != "Not Available":
            result["Work History"] = result["Latest Company"]

        # Handle Competitor Experience - should be blank (empty string) when no match found
        if result["Competitor Experience"] == "Not Available" or not result["Competitor Experience"]:
            # Check work history for competitor names
            result["Competitor Experience"] = check_competitor_experience(result["Work History"], get_planful_competitors())
        elif "no" in result["Competitor Experience"].lower() or "not" in result["Competitor Experience"].lower():
            # If explicitly states no, then make it empty
            result["Competitor Experience"] = ""
        elif not result["Competitor Experience"].lower().startswith("yes"):
            # If doesn't start with "Yes" but has content, check if it's a competitor name
            competitor_found = False
            for competitor in get_planful_competitors():
                if competitor.lower() in result["Competitor Experience"].lower():
                    result["Competitor Experience"] = f"Yes - {competitor}"
                    competitor_found = True
                    break
            if not competitor_found:
                result["Competitor Experience"] = ""
            
        # Clean all text fields
        for field in result:
            result[field] = clean_text(result[field])
            
        # Calculate overall score
        required_experience = 3
        stability_threshold = 2
        
        overall_score, recommendation, individual_scores = calculate_scores(result, required_experience, stability_threshold)
        
        result["Overall Weighted Score"] = str(round(overall_score, 2))
        result["Selection Recommendation"] = recommendation
        
        # Add phone number extraction if resume text is available
        if resume_text:
            result["Phone Number"] = extract_phone_number(resume_text)
        
        return result
    
    except Exception as e:
        st.error(f"Error parsing AI response: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None
